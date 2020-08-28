# make sure the intermine package is installed and the input file and script are in the same folder
# installation command: pip install intermine OR conda install -c bioconda intermine (for anaconda environments)

# input should be an excel file (define file name below) with three columns in this specific order
# (no header required, just plain data): gene name (e.g. adh1; search is always in S. cerevisiae),
# flanking regions (choose from up, down, both, or none if not required),
# length of flanking region (choose from 0.5kb, 1.0kb, 2.0kb, or none if no flanking regions are required)
# do not forget to close the excel file before running the script!

file_name = "xxxx.xlsx"

import csv
from openpyxl import load_workbook
from openpyxl import Workbook

def SGD_sequence_lookup(g, d, l, m):
    from intermine.webservice import Service
    service = Service("https://yeastmine.yeastgenome.org/yeastmine/service")

    query = service.new_query("Gene")
    query.add_constraint("LOOKUP", g, extra_value="S. cerevisiae")

    if d == "none":
        query.select("symbol", "secondaryIdentifier", "length", "sequence.residues")
        m = "Query without flanking regions successful"
    elif d == "up" or d == "down" or d == "both":
        query.select("symbol", "secondaryIdentifier", "length", \
           "flankingRegions.direction", "flankingRegions.sequence.length", "flankingRegions.sequence.residues")
        query.add_constraint("flankingRegions.includeGene", "=", "true")

        if d == "up":
            query.add_constraint("flankingRegions.direction", "=", "upstream")
        elif d == "down":
            query.add_constraint("flankingRegions.direction", "=", "downstream")

        if l == "0.5kb" or l == "1.0kb" or l == "2.0kb":
            query.add_constraint("flankingRegions.distance", "=", l)
            m = "Query with flanking regions successful"
        else:
            m = "Incorrect length of flanking region"
    else:
        m = "Incorrect indication of flanking region"

    reader = csv.reader(query.results(row="csv"), delimiter=",", quotechar='"')
    return reader, m

wb_obj = load_workbook(file_name)
sheet_obj = wb_obj.active
max_row = sheet_obj.max_row

gene_name = []
flank = []
length = []

for i in range(1, max_row+1):
    data1 = sheet_obj.cell(row=i, column=1)
    gene_name.append(data1.value)
    data2 = sheet_obj.cell(row=i, column=2)
    flank.append(data2.value)
    data3 = sheet_obj.cell(row=i, column=3)
    length.append(data3.value)

wb = Workbook()
ws = wb.active

fa_name = file_name.replace(".xlsx","")+".fa"
fasta = open(fa_name,"w")
mes = ""

for i in range(0, max_row):

    if flank[i] == "none":
        ws.append(["Gene", "Identifier", "Gene length", "Sequence"])
    else:
        ws.append(["Gene", "Identifier", "Gene length", "Flanking direction", "Total length", "Sequence"])

    data, message = SGD_sequence_lookup(gene_name[i], flank[i], length[i], mes)
    print(message)

    if message == "Query without flanking regions successful" or message == "Query with flanking regions successful":
        for x in data:
            ws.append(x)
            l = list(x)
            if flank[i] == "none":
                fasta.write(">"+l[0]+" | "+l[1]+" | "+l[2]+" bp \n"+l[3]+"\n")
            else:
                fasta.write(">"+l[0]+" | "+l[1]+" | "+l[4]+" bp | "+l[3]+"\n"+l[5]+"\n")
    else:
        ws.append(["Incorrect query request"])

fasta.close()
wb.save(file_name)